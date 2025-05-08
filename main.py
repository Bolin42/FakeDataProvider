import requests
import datetime
from openpyxl import load_workbook
import os
import time
import json
from outprint import *
import random
import re
import configparser


def put_request(timestamp, time, objectCode, elderAge, elderName, elderSex, heart_max, heart_min, body_temp_max, body_temp_min, sbp_max, sbp_min, dbp_max, dbp_min, blood_glucose_max, blood_glucose_min, pre):
    # 生成随机的 sbp、dbp、hr、thermometer 和 bloodGlucose 值
    info(f"舒张压最小值为{dbp_min}，最大值为{dbp_max}")
    info(f"收缩压最小值为{sbp_min}，最大值为{sbp_max}")
    info(f"心率最小值为{heart_min}，最大值为{heart_max}")
    info(f"体温最小值为{body_temp_min}，最大值为{body_temp_max}，保留{pre}位小数")
    info(f"血糖最小值为{blood_glucose_min}，最大值为{blood_glucose_max}，保留{pre}位小数")
    info("请确认（Y/N）")
    confirm = input()
    if confirm.upper() == 'Y': 
        success("确认上传")
        pass
    else:
        error("取消上传")
        os._exit(0)
    sbp = str(random.randint(sbp_min, sbp_max))
    dbp = str(random.randint(dbp_min, dbp_max))
    hr = str(random.randint(heart_min, heart_max))
    thermometer_ori = str(random.uniform(body_temp_min, body_temp_max))
    thermometer = str(round(float(thermometer_ori), 1))
    bloodGlucose_ori = str(random.uniform(blood_glucose_min, blood_glucose_max))
    bloodGlucose = str(round(float(bloodGlucose_ori), 1))


    url = "https://xn--l6qy95a.fun/baseData/signs/general/add?time="+timestamp

    payload = "{\r\n    \"branchId\": \"23\",\r\n    \"dataTime\": \""+ time +"\",\r\n    \"bmi\": \"\",\r\n    \"sbp\": "+ sbp +",\r\n    \"dbp\": "+ dbp +",\r\n    \"objectCode\": \""+ objectCode + "\",\r\n    \"elderAge\": \""+ elderAge +"\",\r\n    \"elderName\": \""+elderName+"\",\r\n    \"elderSex\": \""+elderSex+"\",\r\n    \"hr\": "+ hr +",\r\n    \"measureNo\": \"\",\r\n    \"thermometer\": "+thermometer+",\r\n    \"remark\": \"\",\r\n    \"bloodGlucose\": "+bloodGlucose+"\r\n}"
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Content-Type': 'application/json;charset=UTF-8',
        'Origin': 'https://xn--l6qy95a.fun',
        'Referer': 'https://xn--l6qy95a.fun/tenantSide/index.html?time='+timestamp+'/',
        'sec-ch-ua-platform': '"Windows"',
        'organId': '23',
        'branchId': '23',
        'sec-ch-ua': '"Chromium";v="134", "Not:A-Brand";v="24", "Microsoft Edge";v="134"',
        'sec-ch-ua-mobile': '?0',
        'Authorization': 'Bearer ' 
    } //should replace authorization with real string 

    response = requests.request("PUT", url, headers=headers, data=payload)

    info(response.text)
    
    if response.status_code == 200:
        success("数据上传成功")
    else:
        error("数据上传失败")

def get_start_date(objectCode, timestamp):
    url = "https://xn--l6qy95a.fun/baseData/signs/general/page?objectCode="+objectCode+"&branchId=23&current=1&size=1&time=" + timestamp

    payload = {}
    headers = {
  'Accept': 'application/json, text/plain, */*',
  'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
  'Referer': 'https://xn--l6qy95a.fun/tenantSide/index.html?time='+timestamp+'/',
  'branchId': '23',
  'organId': '23',
  'sec-ch-ua': '"Chromium";v="134", "Not:A-Brand";v="24", "Microsoft Edge";v="134"',
  'sec-ch-ua-mobile': '?0',
  'sec-ch-ua-platform': '"Windows"',
  'Authorization': 'Bearer'
    } //should replace the authorization with real string 

    response = requests.request("GET", url, headers=headers, data=payload)

    info(response.text)
    if response.status_code == 200:
        response_json = json.loads(response.text)
        try:
            start_date = response_json['data']['data']['records'][0]['dataTime']
            success("目标人员开始时间为"+start_date)
        except IndexError:
            error("目标人员开始时间获取失败")
            start_date = '2024-08-01 10:00:00'
    else:
        error("获取目标人员开始时间失败")
    return start_date
    

def get_stuff_ID(elderNameAndNo, timestamp):
    url = "https://xn--l6qy95a.fun/baseData/data/elderList?elderNameAndNo="+elderNameAndNo+"&size=1&type=all&time="+timestamp

    payload = {}
    headers = {
  'Accept': 'application/json, text/plain, */*',
  'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
  'Referer': 'https://xn--l6qy95a.fun/tenantSide/index.html?time='+timestamp+'/',
  'branchId': '23',
  'organId': '23',
  'sec-ch-ua': '"Chromium";v="134", "Not:A-Brand";v="24", "Microsoft Edge";v="134"',
  'sec-ch-ua-mobile': '?0',
  'sec-ch-ua-platform': '"Windows"',
  'Authorization': 'Bearer '
    } //should replace the authorization with real string 

    response = requests.request("GET", url, headers=headers, data=payload)
    response_json = json.loads(response.text)
    if response.status_code == 200:
        ID = response_json.get('data')['records'][0]['elderId']
        name = response_json.get('data')['records'][0]['elderName']
        sex = response_json.get('data')['records'][0]['sex']
        age = response_json.get('data')['records'][0]['age']
        info(response.text)
        success("目标人员id为"+ID)
        success("目标人员姓名为"+name)
        success("目标人员性别为"+sex)
        success("目标人员年龄为"+age)
    else:
        error("获取目标人员ID等信息失败")
    return ID, name, sex, age

def get_data_from_excel(file_path):
    if os.path.exists(file_path):
        # 加载 Excel 文件
        workbook = load_workbook(file_path)
        # 选择第一个工作表
        sheet = workbook.active
        # 初始化一个列表来存储读取的数据
        data = []
        # 遍历每一行，并将其添加到列表中
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        success("读取数据成功！")
        return data
    else:
        error("文件不存在，请检查路径是否正确。")
        return []

def get_default():
    # 创建一个 ConfigParser 对象
    config = configparser.ConfigParser()

    # 读取配置文件
    config.read('./settings.ini')

    # 获取配置信息并存储到变量中
    try:
        heart_max = config.getint('Vital', 'heart_max')
        heart_min = config.getint('Vital', 'heart_min')
        body_temp_max = config.getfloat('Vital', 'body_temp_max')
        body_temp_min = config.getfloat('Vital', 'body_temp_min')
        sbp_max = config.getint('Vital', 'sbp_max')
        sbp_min = config.getint('Vital', 'sbp_min')
        dbp_max = config.getint('Vital', 'dbp_max')
        dbp_min = config.getint('Vital', 'dbp_min')
        blood_glucose_max = config.getfloat('Vital', 'blood_glucose_max')
        blood_glucose_min = config.getfloat('Vital', 'blood_glucose_min')
        pre = config.getint('Vital', 'pre')
        file_path = config.get('File', 'file_path')
    except configparser.NoSectionError:
        print("Section 'Vital' or 'File' does not exist")
    except configparser.NoOptionError:
        print("Option does not exist in section 'Vital' or 'File'")
    except ValueError:
        print("Value for the option is not an integer or float")

    info(str(heart_max)+str(heart_min)+str(body_temp_max)+str(body_temp_min)+str(sbp_max)+str(sbp_min)+str(dbp_max)+str(dbp_min)+str(blood_glucose_max)+str(blood_glucose_min)+str(pre)+str(file_path))

    return heart_max, heart_min, body_temp_max, body_temp_min, sbp_max, sbp_min, dbp_max, dbp_min, blood_glucose_max, blood_glucose_min, pre, file_path

if __name__ == '__main__':
    heart_max, heart_min, body_temp_max, body_temp_min, sbp_max, sbp_min, dbp_max, dbp_min, blood_glucose_max, blood_glucose_min, pre, file_path = get_default()

    print("===================Powered by Python===========================")
    time.sleep(0.05)
    print(" ##   ##  ######   ##        ####     ####    ##   ##  ######  ")
    time.sleep(0.05)
    print(" ##   ##  ##       ##       ##  ##   ##  ##   ### ###  ##      ")
    time.sleep(0.05)
    print(" ##   ##  ##       ##       ##       ##  ##   #######  ##      ")
    time.sleep(0.05)
    print(" ## # ##  ####     ##       ##       ##  ##   ## # ##  ####    ")
    time.sleep(0.05)
    print(" #######  ##       ##       ##       ##  ##   ##   ##  ##      ")
    time.sleep(0.05)
    print(" ### ###  ##       ##       ##  ##   ##  ##   ##   ##  ##      ")
    time.sleep(0.05)
    print(" ##   ##  ######   ######    ####     ####    ##   ##  ######  ")
    time.sleep(0.05)
    print("===================Dedigned by Bolin===========================")
    time.sleep(0.05)
    print("===================version: v.1.0.0============================")
    time.sleep(0.05)
    print("使用守则：")
    warning("严禁使用本程序及衍生修改版本对服务器进行DOS攻击，否则后果自负！")
    warning("本程序不会检查随机数是否合理，若不确定请使用默认数据！")
    time.sleep(1)
    info("请将包含成员的列表存储到excel文件夹中！")
    info("列表有且只能有一列，即姓名列。程序只会读取第一张表！")
    info("若出现错误，请立即用Ctrl+C退出程序！并立即上报！")
    info("祝您使用顺利！ :-D")
    time.sleep(3)

    #file_path = './excel/excel.xlsx'
    # 读取 Excel 文件
    data = get_data_from_excel(file_path)
    for stuff_ori in data:
        # 正则表达式模式，匹配一个或多个中文字符
        pattern = r'[\u4e00-\u9fff]+'
        # 使用findall方法找到所有匹配的中文字符
        stuff = re.findall(pattern, str(stuff_ori))[0]
        info("当前输入用户："+ str(stuff))
        current_timestamp = str(int(time.time())) #这是post请求中发生的当前时间戳
        info("当前时间戳为:"+ current_timestamp)
        ID, name, sex, age = get_stuff_ID(str(stuff), current_timestamp)
        start_date = get_start_date(ID, current_timestamp) #这是get请求中发生的最近时间戳
        # 首先，将 startTimeDate 转换为 datetime 对象
        startTimeDate = datetime.strptime(start_date, "%Y-%m-%d %H:%M:%S")
        # 然后，将 datetime 对象转换为时间戳
        startTimeStamp = int(time.mktime(startTimeDate.timetuple()))

        while startTimeStamp <= 1740672000:
            startTimeStamp += 604800
            date_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(startTimeStamp))
            put_request(current_timestamp, date_str, ID, age, name, sex)
            